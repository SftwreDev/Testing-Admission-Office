from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import CampusForm, CollegeForm, MajorForm, CourseForm, StudentProfileForm, StudentPreviousSchoolForm, UpdateStudentPreviousSchoolForm
from .models import Campus, College, Course, Major, StudentProfile, PreviousSchool
from .filters import StudentProfileFilter

from repository.forms import FileUploadForm
from repository.models import FileUpload

from django.http import HttpResponse
import xlwt
import csv
import datetime
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

@login_required
def student_profile(request):
    template = 'students/student_profile.html'
    student_list = StudentProfile.objects.all()
    student_form = StudentProfileForm(request.POST or None)
    filter_qs = StudentProfileFilter(request.GET, queryset=student_list)

    if request.method == 'POST':
        student_form = StudentProfileForm(request.POST)
        if student_form.is_valid():
            student_form.save()
            return redirect('students:student_profile')
        else:
            student_form = StudentProfileForm(request.POST or None)

    context = {
        'student_form' : student_form,
        'student_list' : student_list,
        'filter_qs' : filter_qs ,
    }
    return render(request, template, context)
            

@login_required
def students_detail_view(request, pk):
    template_name = "students/students_detailed_view.html"
    form = FileUploadForm(request.POST or None, request.FILES or None)

    student = StudentProfile.objects.get(id=pk)
    files = FileUpload.objects.filter(student=student)
    update_form = StudentProfileForm(request.POST or None , instance=student)

   
    if form.is_valid():
        files = form.save(commit=False)
        files.student = student
        files.save()
        messages.success(request, "Successfully adding a document")
        return redirect('students:students_detail_view', pk)
    else:
        print("error")
        form = FileUploadForm(request.POST or None)

    if update_form.is_valid():
        update_form.save()
        messages.success(request, "You have successfully update an information")
        return redirect("students:students_detail_view", pk)
    

    context = {
        'student' : student,
        'form' : form ,
        'files' : files,
        'update_form' : update_form,
    }

    return render(request, template_name, context)


@login_required
def students_previous_school_detail_view(request, pk):
    template_name = "students/students_previous_school_detailed_view.html"
   
    try:
        student = StudentProfile.objects.get(id=pk)
        previous_school = PreviousSchool.objects.get(student=student)
        
        form = StudentPreviousSchoolForm(request.POST or None)
        update_form = UpdateStudentPreviousSchoolForm(request.POST or None, instance=previous_school)
        if request.method == 'POST':
            form = StudentPreviousSchoolForm(request.POST or None)
            update_form = StudentPreviousSchoolForm(request.POST or None, instance=previous_school)
            if form.is_valid():
                previous = form.save(commit=False)
                previous.student = student
                previous.save()
                messages.info(request, "Success !")
                return redirect("students:students_previous_school_detail_view", pk)
            else:
                messages.error(request, "Invalid input of data !")
                form = StudentPreviousSchoolForm(request.POST or None)

        if request.method == 'POST':
            update_form = UpdateStudentPreviousSchoolForm(request.POST or None, instance=previous_school)
            if update_form.is_valid():
                update_form.save()
                messages.success(request, "You have successfully update an information")
                return redirect("students:students_previous_school_detail_view", pk)
            else:
                messages.success(request, "Invalid input!")
                return redirect("students:students_previous_school_detail_view", pk)

        context = {
            'student' : student,
            'form' : form,
            'previous_school' :previous_school,
            'update_form' : update_form
        }
            
    except:
        student = StudentProfile.objects.get(id=pk)
        
        form = StudentPreviousSchoolForm(request.POST or None)

        if request.method == 'POST':
            form = StudentPreviousSchoolForm(request.POST or None)
            if form.is_valid():
                previous = form.save(commit=False)
                previous.student = student
                previous.save()
                messages.info(request, "Success !")
                return redirect("students:students_previous_school_detail_view", pk)
            else:
                messages.error(request, "Invalid input of data !")
                form = StudentPreviousSchoolForm(request.POST or None)
        context = {
            'student' : student,
            'form' : form,
            
        }

    return render(request, template_name, context)


@login_required
def students_previous_school_update_view(request, pk):
    template_name = "students/students_previous_school_update_view.html"
   
    try:
        student = StudentProfile.objects.get(id=pk)
        previous_school = PreviousSchool.objects.get(student=student)
        
        update_form = UpdateStudentPreviousSchoolForm(request.POST or None, instance=previous_school)
        
        if request.method == 'POST':
            update_form = UpdateStudentPreviousSchoolForm(request.POST or None, instance=previous_school)
            if update_form.is_valid():
                form = update_form.save(commit=False)
                form.user = request.user
                form.save()
                messages.success(request, "You have successfully update an information")
                return redirect("students:students_previous_school_detail_view", pk)
            else:
                messages.success(request, "Invalid input!")
                return redirect("students:students_previous_school_detail_view", pk)

        context = {
            'student' : student,
            'previous_school' :previous_school,
            'update_form' : update_form
        }
            
    except:
        student = StudentProfile.objects.get(id=pk)
        
        form = StudentPreviousSchoolForm(request.POST or None)

        if request.method == 'POST':
            form = StudentPreviousSchoolForm(request.POST or None)
            if form.is_valid():
                previous = form.save(commit=False)
                previous.student = student
                previous.save()
                messages.info(request, "Success !")
                return redirect("students:students_previous_school_detail_view", pk)
            else:
                messages.error(request, "Invalid input of data !")
                form = StudentPreviousSchoolForm(request.POST or None)
        context = {
            'student' : student,
            'form' : form,
            
        }

    return render(request, template_name, context)

def delete_students(request, pk):

    obj = StudentProfile.objects.get(id=pk)
    obj.delete()

    return redirect('students:student_profile')

def delete_previous_school(request, pk):
    template_name = 'students/delete_previous_school.html'
    
    obj = PreviousSchool.objects.get(id=pk)
    obj.delete()

    return redirect("students:student_profile")


def export_students(request):
  
    category_queryset = StudentProfile.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-student-information.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Student Profile'

    


    # Define the column titles and widths
    columns = [
        ('ID'),
        ('Last Name'),
        ('First Name'),
        ('Middle Name'),
        ('College'),
        ('Course'),
        ('Major'),
        ('Year Level'),
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        wrapped_alignment = Alignment(vertical='top', wrap_text=True)
        cell.alignment = wrapped_alignment
    # Iterate through all lists
    for list in category_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            list.pk,
            list.l_name,
            list.f_name,
            list.m_name,
            list.college.college,
            list.course.course,
            list.major.major,
            list.yearlevel,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def settings(request):
    template_name = "settings.html"

    college = College.objects.all()
    course = Course.objects.all()
    major = Major.objects.all()
    campus = Campus.objects.all()
    
    campusform = CampusForm(request.POST or None)
    collegeform = CollegeForm(request.POST or None)
    majorform = MajorForm(request.POST or None)
    courseform = CourseForm(request.POST or None)

    try:
        if request.method == 'POST':
            campusform = CampusForm(request.POST or None)
            collegeform = CollegeForm(request.POST or None)
            majorform = MajorForm(request.POST or None)
            courseform = CourseForm(request.POST or None)
            if campusform.is_valid():
                campusform.save()
                return redirect('students:settings')
            else:
                campusform = CampusForm(request.POST or None)
            if collegeform.is_valid():
                collegeform.save()
                return redirect('students:settings')
            else:
                collegeform = CollegeForm(request.POST or None)
            if majorform.is_valid():
                majorform.save()
                return redirect('students:settings')
            else:
                majorform = MajorForm(request.POST or None)
            if courseform.is_valid():
                courseform.save()
                return redirect('students:settings')
            else:
                courseform = CourseForm(request.POST or None)
    except:
        return redirect('students:settings')
            
    context = {
        'college' : college,
        'course' : course,
        'major' : major,
        'campus' : campus,
        'collegeform' : collegeform,
        'majorform' : majorform,
        'courseform' : courseform,
        'campusform' : campusform
    }

    return render(request, template_name, context)


def settings_update_form(request, pk):
    template_name = "settings_update.html"

    college = College.objects.get(id=pk)

    collegeform = CollegeForm(request.POST or None, instance=college)


    
    try:
        if request.method == 'POST':
            
            collegeform = CollegeForm(request.POST or None, instance=college)
    
           
            if collegeform.is_valid():
                collegeform.save()
                return redirect('students:settings')
            else:
                collegeform = CollegeForm(request.POST or None)
                return redirect('students:settings')
    except:
        return redirect('students:settings')
            
    context = {

        'collegeform' : collegeform,
   
    }

    return render(request, template_name, context)

def settings_update_course(request, pk):
    template_name = "settings_update_course.html"

    course = Course.objects.get(id=pk)

    courseform = CourseForm(request.POST or None, instance=course)

    
    try:
        if request.method == 'POST':
          
            courseform = CourseForm(request.POST or None, instance=course)
            if courseform.is_valid():
                courseform.save()
                return redirect('students:settings')
            else:
                courseform = CourseForm(request.POST or None)
    except:
        return redirect('students:settings')
            
    context = {

        'courseform' : courseform,

    }

    return render(request, template_name, context)
def settings_update_majors(request, pk):
    template_name = "settings_update_majors.html"


    major = Major.objects.get(id=pk)

    majorform = MajorForm(request.POST or None, instance=major)


    
    try:
        if request.method == 'POST':

            majorform = MajorForm(request.POST or None, instance=major)

            if majorform.is_valid():
                majorform.save()
                return redirect('students:settings')
            else:
                majorform = MajorForm(request.POST or None)

    except:
        return redirect('students:settings')
            
    context = {

        'majorform' : majorform,

    }

    return render(request, template_name, context)

def settings_update_campus(request, pk):
    template_name = "settings_update_campus.html"


    campus = Campus.objects.get(id=pk)
    
    campusform = CampusForm(request.POST or None, instance=campus)

    
    try:
        if request.method == 'POST':
            campusform = CampusForm(request.POST or None, instance=campus)

            if campusform.is_valid():
                campusform.save()
                return redirect('students:settings')
            else:
                campusform = CampusForm(request.POST or None)
            
    except:
        return redirect('students:settings')
            
    context = {

        'campusform' : campusform
    }

    return render(request, template_name, context)


def delete_college(request, pk):

    college = College.objects.get(id=pk)
    college.delete()

    return redirect('students:settings')

def delete_course(request, pk):

    course = Course.objects.get(id=pk)
    course.delete()

    return redirect('students:settings')

def delete_majors(request, pk):

    major = Major.objects.get(id=pk)
    major.delete()

    return redirect('students:settings')

def delete_campus(request, pk):

    campus = Campus.objects.get(id=pk)
    campus.delete()

    return redirect('students:settings')


    